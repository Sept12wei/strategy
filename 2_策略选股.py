from SMA_strategy import *
from datetime import datetime
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
from openpyxl import Workbook

# 获取当前日期并格式化为字符串，格式为YYYY-MM-DD
current_date = datetime.now().strftime('%Y-%m-%d')
# 创建一个timedelta对象，表示一天的时间
one_day = timedelta(days=1)
# 计算当前日期的下一天
next_day = datetime.strptime(current_date, '%Y-%m-%d') + one_day
next_day_str = next_day.strftime('%Y-%m-%d')

risk_free_rate = 0.01
start_date = "2015-01-01"
end_date = next_day_str
print(current_date)
file_path = './竞价交易策略.xlsx'
output_file_path = './竞价交易.xlsx'

df = pd.read_excel(file_path, engine='openpyxl')

def create_sheet(output_file_path):
    """
    创建表单
    """
    # 确保Excel文件存在，如果不存在则创建
    if not os.path.exists(output_file_path):
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            # 创建三个空的工作表
            for sheet_name in ['1d', '1wk', '1mo']:
                pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)

    # 加载现有的工作簿
    wb = load_workbook(output_file_path)

    # 检查特定的工作表是否存在，如果不存在则创建
    for sheet_name in ['1d', '1wk', '1mo']:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
    # 保存工作簿以应用更改
    wb.save(filename=output_file_path)
    wb.close()


def create_empty_excel_file(output_file_path):
    # 创建一个新的工作簿对象
    wb = Workbook()

    # 保存并关闭工作簿，这将创建一个没有任何工作表的Excel文件
    wb.save(filename=output_file_path)
    wb.close()


def SMA(short, long, frequency):
    """
    均线策略
    """
    # 创建一个空的DataFrame来存储结果
    results_df = pd.DataFrame(
        columns=['股票代码', '胜率', '盈亏比', '最大回撤', '夏普比率', '行业', 'position', '年化波动率'])

    # 遍历第二列中的股票代码
    for stock_code in df['代码']:
        old_code = stock_code
        if stock_code.startswith('SH'):
            stock_code = f"{stock_code[2:]}.{'SS'}"
        elif stock_code.startswith('SZ'):
            stock_code = f"{stock_code[2:]}.{'SZ'}"
        print(stock_code)

        data = plot_stock_strategy(stock_code, start_date, end_date, short, long, frequency)
        print(f'data-->{data}')
        # 将股票代码和指标数据作为一个新行添加到结果DataFrame中
        results_df = results_df.append({
            '股票代码': stock_code,
            '胜率': calculate_win_rate(data),
            '盈亏比': calculate_ror(data),
            '最大回撤': calculate_max_drawdown(data),
            '夏普比率': calculate_sharpe_ratio(data, risk_free_rate),
            '年化波动率': calculate_volatility(data),
            '行业': df.loc[df['代码'] == f'{old_code}', '所属行业'].values[0],
            'position': data['position'].iloc[-1]
        }, ignore_index=True)

    writer = pd.ExcelWriter(output_file_path, engine='openpyxl')
    writer.book = load_workbook(output_file_path)  # 加载现有的工作簿
    # 检查特定名称的工作表是否存在，如果存在，则删除它
    if frequency in writer.book.sheetnames:
        del writer.book[frequency]
    # 将DataFrame写入指定的工作表，如果工作表存在，to_excel会追加数据
    results_df.to_excel(writer, sheet_name=frequency, index=False, startrow=0, header=True)

    writer.save()
    writer.close()

    print(f'指标数据已写入 {output_file_path} 的 {frequency} 工作表')
    return


if __name__ == '__main__':
    create_sheet(output_file_path)
    # create_empty_excel_file(output_file_path)
    # 计算日均线策略
    SMA(5, 10, '1d')
    # 计算周均线策略
    SMA(5, 10, '1wk')
    # 计算月均线策略
    SMA(1, 3, '1mo')





